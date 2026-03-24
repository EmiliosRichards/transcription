"""
00_prepare_inputs.py
Convert the two Excel source files into machine-readable formats:
  - Kundenzitate und Antworten 3.xlsx  →  config/taxonomy.json
  - Dexter Jagdhütte.xlsx              →  config/referenzen.csv

Usage:
    python scripts/00_prepare_inputs.py
    python scripts/00_prepare_inputs.py --excel-dir .   # if Excel files are elsewhere
"""

import argparse
import csv
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CONFIG_DIR = ROOT / "config"


# ---------------------------------------------------------------------------
# Taxonomy (Kundenzitate und Antworten 3.xlsx)
# ---------------------------------------------------------------------------

def _clean(text: str | None) -> str:
    """Strip and normalise whitespace."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def _split_quotes(cell_value: str | None) -> list[str]:
    """DM sheet packs multiple quotes into one cell separated by newlines."""
    if not cell_value:
        return []
    return [q.strip() for q in str(cell_value).split("\n") if q.strip()]


def _is_exclude(answer: str) -> bool:
    """Check if the answer template indicates 'raus' (exclude from calling)."""
    low = answer.lower().strip()
    return low.startswith("raus")


def parse_gk_sheet(ws) -> list[dict]:
    categories = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row[0]:
            continue
        label = _clean(row[0])
        quotes = [_clean(row[i]) for i in range(1, 5) if row[i]]
        answer = _clean(row[5]) if len(row) > 5 else ""
        action = "exclude" if _is_exclude(answer) else "pitch"
        categories.append({
            "id": f"GK{idx:02d}",
            "role": "GK",
            "label": label,
            "quotes": quotes,
            "answer_template": answer,
            "action": action,
        })
    return categories


def parse_dm_sheet(ws) -> list[dict]:
    categories = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row[0]:
            continue
        label = _clean(row[0])
        quotes = _split_quotes(row[1])
        answer_individual = _clean(row[2]) if len(row) > 2 else ""
        answer_generic = _clean(row[3]) if len(row) > 3 else ""
        action = "exclude" if _is_exclude(answer_individual) else "pitch"
        categories.append({
            "id": f"DM{idx:02d}",
            "role": "DM",
            "label": label,
            "quotes": quotes,
            "answer_template": answer_individual,
            "generic_followup": answer_generic,
            "action": action,
        })
    return categories


def build_taxonomy(excel_path: Path) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    gk = parse_gk_sheet(wb["GK_Gatekeeper"])
    dm = parse_dm_sheet(wb["DM_Entscheider"])
    return {
        "gatekeeper": gk,
        "decision_maker": dm,
        "meta": {
            "source": excel_path.name,
            "gk_count": len(gk),
            "dm_count": len(dm),
        },
    }


# ---------------------------------------------------------------------------
# Referenzen (Dexter Jagdhütte.xlsx → Referenzen_200226 sheet)
# ---------------------------------------------------------------------------

def build_referenzen(excel_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Referenzen_200226"]

    # Find header row (contains "Träger" or similar)
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if any("tr" in v and "ger" in v for v in vals if v):
            header_row = row_idx
            break

    if header_row is None:
        header_row = 4

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        traeger = _clean(row[0]) if row[0] else ""
        einrichtung = _clean(row[1]) if len(row) > 1 and row[1] else ""
        ort = _clean(row[2]) if len(row) > 2 and row[2] else ""
        plz = _clean(row[3]) if len(row) > 3 and row[3] else ""
        system = _clean(row[4]) if len(row) > 4 and row[4] else ""

        # Skip empty rows
        if not einrichtung and not traeger:
            continue

        rows.append({
            "traeger": traeger,
            "einrichtung": einrichtung,
            "ort": ort,
            "plz": plz,
            "system": system,
        })
    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Convert Excel inputs to JSON/CSV")
    parser.add_argument("--excel-dir", type=Path, default=ROOT,
                        help="Directory containing the Excel files")
    args = parser.parse_args()

    CONFIG_DIR.mkdir(parents=True, exist_ok=True)

    # --- Taxonomy ---
    tax_path = args.excel_dir / "Kundenzitate und Antworten 3.xlsx"
    if not tax_path.exists():
        print(f"ERROR: {tax_path} not found")
        return
    taxonomy = build_taxonomy(tax_path)
    out_tax = CONFIG_DIR / "taxonomy.json"
    with open(out_tax, "w", encoding="utf-8") as f:
        json.dump(taxonomy, f, ensure_ascii=False, indent=2)
    print(f"Taxonomy: {taxonomy['meta']['gk_count']} GK + {taxonomy['meta']['dm_count']} DM categories -> {out_tax}")

    # --- Referenzen ---
    ref_path = args.excel_dir / "Dexter Jagdhütte.xlsx"
    if not ref_path.exists():
        print(f"ERROR: {ref_path} not found")
        return
    refs = build_referenzen(ref_path)
    out_ref = CONFIG_DIR / "referenzen.csv"
    with open(out_ref, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["traeger", "einrichtung", "ort", "plz", "system"])
        writer.writeheader()
        writer.writerows(refs)
    print(f"Referenzen: {len(refs)} facilities -> {out_ref}")


if __name__ == "__main__":
    main()
