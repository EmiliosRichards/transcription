import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

try:
    from dotenv import load_dotenv, find_dotenv  # type: ignore
    load_dotenv(find_dotenv(usecwd=True))
except Exception:
    pass


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Extract fixed GK/DM categories + quotes from Excel into JSON.")
    p.add_argument("--xlsx", required=True, help="Path to GK_DM_Kategorien und Kundenzitate.xlsx")
    p.add_argument("--out", required=True, help="Path to write categories.json")
    p.add_argument("--gk-sheet", default="GK_Gatekeeper", help="Worksheet name for Gatekeeper categories")
    p.add_argument("--dm-sheet", default="DM_Entscheider", help="Worksheet name for Decision Maker categories")
    return p.parse_args()


def _read_sheet(ws) -> List[Dict[str, Any]]:
    # Expected columns:
    # A: Kategorie, B-E: Kundenzitat 1..4
    rows: List[Dict[str, Any]] = []
    for r in range(2, (ws.max_row or 0) + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label_s = str(label).strip()
        if not label_s:
            continue
        quotes: List[str] = []
        for c in range(2, 6):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                quotes.append(s)
        rows.append({"label": label_s, "quotes": quotes})
    return rows


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise SystemExit(f"openpyxl is required. Install it first. Import error: {e}")

    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if args.gk_sheet not in wb.sheetnames:
        raise SystemExit(f"GK sheet '{args.gk_sheet}' not found. Sheets: {wb.sheetnames}")
    if args.dm_sheet not in wb.sheetnames:
        raise SystemExit(f"DM sheet '{args.dm_sheet}' not found. Sheets: {wb.sheetnames}")

    gk = _read_sheet(wb[args.gk_sheet])
    dm = _read_sheet(wb[args.dm_sheet])

    # Assign stable IDs
    gk_items = [{"id": f"GK{i:02d}", **item} for i, item in enumerate(gk, start=1)]
    dm_items = [{"id": f"DM{i:02d}", **item} for i, item in enumerate(dm, start=1)]

    payload = {
        "gatekeeper": gk_items,
        "decision_maker": dm_items,
        "meta": {
            "source_xlsx": str(xlsx_path),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "gk_sheet": args.gk_sheet,
            "dm_sheet": args.dm_sheet,
            "gk_count": len(gk_items),
            "dm_count": len(dm_items),
            "note": "This taxonomy is fixed. Classifier must choose from these labels or Other/Unclear.",
        },
    }

    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print("Wrote:", out_path)
    print(f"GK categories: {len(gk_items)} | DM categories: {len(dm_items)}")


if __name__ == "__main__":
    main()

